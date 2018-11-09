import time
import urllib
import numpy as np
import cv2
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import csv
from StringIO import StringIO
from keras_retinanet import models
from keras_retinanet.utils.image import read_image_bgr, preprocess_image, resize_image
from keras_retinanet.utils.visualization import draw_box, draw_caption
from keras_retinanet.utils.colors import label_color

# xlsx_path = '/Users/cashify/Downloads/ml_data.xlsx'
# wb = load_workbook(xlsx_path)
# ws = wb.worksheets[0]


# ### XLSX TO CSV
# def iter_rows(worksheet):
#     l = []
#     for row in worksheet.iter_rows():
#         list1 = [cell.value for cell in row]
#         l.append(list1)
#     return l
#
# row_list = iter_rows(ws)
# del row_list[0]
#
# print row_list
#
# f = open('id_proof/data.csv', 'wb')
# csv.writer(f).writerows(row_list)


#XLSX TO LABEL CSV
# label_list = []
# for i in range(ws.max_row):
#     if i <= 1:
#         continue
#     cell_value = ws[i][ws.max_column-1].value
#     if cell_value not in label_list:
#         label_list.append(cell_value)
#
# label_list = [[label_list[i], i] for i in range(len(label_list))]
# print label_list
#
# f = open('id_proof/labels.csv', 'w')
# csv.writer(f).writerows(label_list)

#DOWNLOAD PICS AND MODIFY XLSX
# for i in range(2, ws.max_row+1):
#     try:
#         filename = ws[i][0].value.split('/')[-1].split('.')[-2] + '.jpg'
#         urllib.urlretrieve(ws[i][0].value, 'id_proof/' + filename)
#     except Exception:
#         ws[i][0].value = 'fail'
#         print 'fail'
#         continue
#
#     ws[i][0].value = filename
#
# wb.save(xlsx_path)

# TESTING
model = models.load_model('dogs/inf_model_2.h5', backbone_name='resnet50')
# load image
image = read_image_bgr('id_proof/0AlZx0YI-image.jpg')

# copy to draw on
draw = image.copy()
draw = cv2.cvtColor(draw, cv2.COLOR_BGR2RGB)

# preprocess image for network
image = preprocess_image(image)
image, scale = resize_image(image)

# process image
start = time.time()
boxes, scores, labels = model.predict_on_batch(np.expand_dims(image, axis=0))
print scores
print("processing time: ", time.time() - start)

# correct for image scale
boxes /= scale

# visualize detections
for box, score, label in zip(boxes[0], scores[0], labels[0]):
    # scores are sorted so we can break
    if score < 0.5:
        break

    color = label_color(label)

    b = box.astype(int)
    draw_box(draw, b, color=color)

    caption = "{} {:.3f}".format(label, score)
    draw_caption(draw, b, caption)

plt.figure(figsize=(15, 15))
plt.axis('off')
plt.imshow(draw)
plt.savefig('abc.jpg')

